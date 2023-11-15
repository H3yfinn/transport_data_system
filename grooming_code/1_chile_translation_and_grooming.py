#%%
#note this is an old file thta was used, for some data that didnt end up being what i wanted. so keeping it here just in case it comes in useful once we find the right data
#%%
#take in data files. they all have around about the same structure:
import os
import pandas as pd
import numpy as np
import plotly.express as px
import plotly.graph_objects as go
import re
import datetime
os.chdir(re.split('transport_data_system', os.getcwd())[0]+'/transport_data_system')
#%%

from openpyxl import load_workbook
import pandas as pd
from googletrans import Translator

# Initialize the translator
translator = Translator()

def translate_text(text):
    if text is None or pd.isna(text):
        return text
    elif isinstance(text, str):
        try:
            return translator.translate(text, src='es', dest='en').text
        except Exception as e:
            print(f"Error translating text: {text}. Error: {e}")
            return text  # Return the original text if translation fails
    else:
        return text  # For any other non-string type, return as is

# Load the Excel workbook
file_path = 'input_data/Chile/cap20001_7.xlsx'
wb = load_workbook(file_path)

# List the sheet names
sheet_names = wb.sheetnames
#%%

# Create a new Excel writer object for the translated workbook
output_file_path = 'input_data/Chile/road_transport_statistics_2020_english.xlsx'
with pd.ExcelWriter(output_file_path, engine='openpyxl') as writer:
    
    # Translate each sheet
    for sheet_name in sheet_names:
        df = pd.read_excel(file_path, sheet_name=sheet_name)
        df_translated = df.applymap(translate_text)
        df_translated.to_excel(writer, sheet_name=sheet_name, index=False)
    
    # Get the workbook and worksheet for adding the link
    workbook  = writer.book
    
    # Set visibility and activation for the first sheet
    first_sheet = workbook[sheet_names[0]]
    first_sheet.sheet_state = 'visible'  # Make sure it's visible
    workbook.active = first_sheet  # Make it the active sheet
    
    # Add the source link to the first sheet at cell A1
    first_sheet['A1'] = 'Source Link: https://www.ine.gob.cl/estadisticas/economia/transporte-y-comunicaciones/estructura-del-transporte-por-carretera'
#%%